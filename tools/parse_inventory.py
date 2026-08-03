"""
Regenerates lib/data/seed_search_products.dart directly from your two
master inventory registers.

USAGE
-----
    pip install openpyxl
    python3 parse_inventory.py --admin cda_admin.xlsx --ops cda_ops.xlsx --out seed_search_products.dart

All three flags are optional — if omitted they default to looking for
cda_admin.xlsx / cda_ops.xlsx in the current folder, and writing
seed_search_products.dart in the current folder. Copy that output file
over lib/data/seed_search_products.dart in your Flutter project when done.

WHAT IT PARSES
--------------
Parses cda_admin.xlsx (Branch 1) and cda_ops.xlsx (Branch 2) into a flat,
unified list of product records:

    {branch, room, row, tray, name, quantity, notes}

Layout of the source sheets (hand-maintained inventory registers):
  - Each "room" is one or two spreadsheet columns: an item-name column,
    optionally followed by a QTY column.
  - Inside a room's column, some cells are not products at all — they are
    inline section markers the staff typed directly into the item list to
    mark a physical sub-location, e.g. "TRAY-4", "DRAW 1", "UPPER ROW".
    Every product listed under a marker (until the next marker) belongs to
    that tray/sub-location.
  - When a room's column has a paired QTY column, a cell with a name but
    an EMPTY qty is treated as a marker/section-label, not a product (real
    products in these sheets are always given a quantity). When a room's
    column has no paired QTY column at all (a handful of columns), every
    row is a product and a trailing "-<digits>" on the name is read as
    the quantity if present (falls back to 1).
"""
import argparse
import re
import json
import openpyxl

TRAY_RE = re.compile(r'^\s*(TRAY|DRAW)\b', re.IGNORECASE)
ROW_WORD_RE = re.compile(r'\bROW\b', re.IGNORECASE)
TRAILING_QTY_RE = re.compile(r'-(\d+)\s*$')
LEADING_INT_RE = re.compile(r'^\s*(\d+)')

# Section-label-only text that shows up with no qty and should never become
# a product of its own (it's a sub-heading like "BATTERIES" grouping the
# real, coded items listed right under it).
LABEL_BLACKLIST = {
    'BATTERIES', 'BATTERY', 'REMOTE CONTROLLER', 'REMOTE CONTROLLERS',
    'RIGHT SIDE', 'LEFT SIDE', 'COMPLETED DOCUMENTS', 'ITEMS', 'STATIONARIES',
    'BROUCHERS', 'CATLOGS',
}

QTY_LABELS = {'QTY', 'QUANTITY', 'QUANTITES', 'QUANTITIES'}
SKIP_LABELS = {'DATE IN', 'DATE OUT'}


_ACRONYMS = {
    'RC', 'TC', 'IC', 'AC', 'DC', 'LED', 'USB', 'GPS', 'DJI', 'FPV', 'MD',
    'PVC', 'XT', 'VTX', 'ESC', 'FC', 'GHZ', 'MHZ', 'CDA', 'RPTO', 'MR',
    'GK', 'WIFI', 'CCTV', 'UPS', 'SOP', 'BMS', 'CPU',
}
_LOWER_WORDS = {'and', 'or', 'of', 'in', 'to', 'on', 'at', 'the', 'with', 'for'}
_WORD_RE = re.compile(r'^([^A-Za-z0-9]*)([A-Za-z0-9.\-]*)([^A-Za-z0-9]*)$')


def titlecase(s: str) -> str:
    s = re.sub(r'\s+', ' ', s.strip())
    if not s:
        return s
    words = s.split(' ')
    out = []
    for i, w in enumerate(words):
        m = _WORD_RE.match(w)
        lead, body, trail = m.groups() if m else ('', w, '')
        if not body:
            out.append(w)
            continue
        bu = body.upper()
        if bu in _ACRONYMS or any(c.isdigit() for c in body):
            newbody = bu
        elif body.lower() in _LOWER_WORDS and i != 0:
            newbody = body.lower()
        else:
            newbody = body[:1].upper() + body[1:].lower()
        out.append(lead + newbody + trail)
    return ' '.join(out)


_HEADER_FIXES = {
    'SIMULTION TRANSMITTER': 'SIMULATION TRANSMITTER',
    'REST ROOM THING': 'RESTROOM THINGS',
}


def _fix_header(h: str) -> str:
    up = h.upper()
    for wrong, right in _HEADER_FIXES.items():
        if wrong in up:
            idx = up.index(wrong)
            h = h[:idx] + right + h[idx + len(wrong):]
            up = h.upper()
    return h


def clean_room_name(header: str) -> tuple[str, str | None]:
    header = _fix_header(header)
    """Splits a header like 'SIMULTION TRANSMITTER(ROW-1)' into
    (room, row) = ('Simulation Transmitter', 'Row 1'). Headers that ARE a
    row on their own (e.g. 'ROW-2', 'ROW 6 REMOTE CONTROLLERS') map to
    room == row.
    """
    h = header.strip()
    m = re.match(r'^(.*?)\(([^)]*ROW[^)]*)\)\s*$', h, re.IGNORECASE)
    if m:
        room = titlecase(m.group(1))
        row_raw = m.group(2)
        rn = re.search(r'(\d+)', row_raw)
        row = f'Row {rn.group(1)}' if rn else titlecase(row_raw)
        return room, row
    if re.match(r'^ROW[\s\-]*\d', h, re.IGNORECASE):
        rn = re.search(r'(\d+)', h)
        extra = h[rn.end():].strip(' -') if rn else ''
        row = f'Row {rn.group(1)}' if rn else titlecase(h)
        room = titlecase(f'Row {rn.group(1)} {extra}'.strip()) if extra else row
        return room, row
    return titlecase(h), None


def parse_qty(cell, name_had_suffix_qty=None):
    if cell is None:
        return None, None
    if isinstance(cell, (int, float)):
        return int(cell), None
    s = str(cell).strip()
    if s == '':
        return None, None
    if s.upper() == 'OUT':
        return 0, 'Marked OUT in source sheet'
    m = LEADING_INT_RE.match(s)
    if m:
        rest = s[m.end():].strip()
        return int(m.group(1)), (rest if rest else None)
    return 1, s  # non-numeric qty text (e.g. "1BOX" missed, "FEW") -> qty 1, note raw


def clean_item_name(raw: str) -> tuple[str, int | None]:
    """Strips a trailing '-<digits>' quantity off an item name typed like
    'STUDENT CHAIR-8'. Returns (clean_name, qty_or_None)."""
    s = raw.strip()
    m = TRAILING_QTY_RE.search(s)
    if m:
        qty = int(m.group(1))
        name = s[:m.start()].strip(' -')
        if name:
            return name, qty
    return s, None


def is_marker(name: str, qty_cell_present: bool, qty_cell_value) -> bool:
    upper = name.strip().upper()
    if TRAY_RE.match(name) or ROW_WORD_RE.search(name):
        return True
    if qty_cell_present and (qty_cell_value is None or str(qty_cell_value).strip() == ''):
        if upper in LABEL_BLACKLIST:
            return True
        # A short, all-caps, digit-free label with no qty in a paired
        # column is almost certainly a sub-heading, not a real product.
        if upper.isupper() and not any(c.isdigit() for c in upper) and len(upper.split()) <= 3:
            return True
    return False


def marker_to_tray(name: str) -> str:
    return titlecase(name)


def classify_columns(header_row):
    """Returns list of (name_col_idx, qty_col_idx_or_None, header_text)."""
    cols = []
    n = len(header_row)
    i = 0
    while i < n:
        h = header_row[i]
        if h is None:
            i += 1
            continue
        htext = str(h).strip()
        if htext.upper() in QTY_LABELS or htext.upper() in SKIP_LABELS or htext.upper() == 'BRANCH':
            i += 1
            continue
        nxt = header_row[i + 1] if i + 1 < n else None
        nxt_text = str(nxt).strip().upper() if nxt is not None else ''
        if nxt_text in QTY_LABELS:
            cols.append((i, i + 1, htext))
            i += 2
        else:
            cols.append((i, None, htext))
            i += 1
    return cols


def parse_paired_or_unpaired_sheet(rows, header_row_idx, branch, extra_room_prefix=None,
                                    branch_col=None, sheet_label=''):
    header_row = rows[header_row_idx]
    cols = classify_columns(header_row)
    records = []
    for name_idx, qty_idx, header_text in cols:
        room, row = clean_room_name(header_text)
        if extra_room_prefix:
            room = f'{extra_room_prefix} · {room}' if room.upper() != extra_room_prefix.upper() else room
        current_tray = None
        for r in rows[header_row_idx + 1:]:
            if name_idx >= len(r):
                continue
            raw = r[name_idx]
            if raw is None:
                continue
            name = str(raw).strip()
            if not name:
                continue
            qty_present = qty_idx is not None
            qty_cell = r[qty_idx] if (qty_idx is not None and qty_idx < len(r)) else None
            if is_marker(name, qty_present, qty_cell):
                current_tray = marker_to_tray(name)
                continue
            note = None
            if qty_present:
                qty, note = parse_qty(qty_cell)
                if qty is None:
                    qty = 1
                clean_name = name
                m = TRAILING_QTY_RE.search(clean_name)
                if m and int(m.group(1)) == qty:
                    stripped = clean_name[:m.start()].strip(' -')
                    if stripped:
                        clean_name = stripped
            else:
                clean_name, suffix_qty = clean_item_name(name)
                qty = suffix_qty if suffix_qty is not None else 1
            if not clean_name or clean_name.isdigit():
                continue
            this_branch = branch
            if branch_col is not None and branch_col < len(r) and r[branch_col]:
                btext = str(r[branch_col]).strip().upper()
                this_branch = 'Branch 1' if 'BRANCH 1' in btext else ('Branch 2' if 'BRANCH 2' in btext else branch)
            records.append({
                'branch': this_branch,
                'room': room,
                'row': row,
                'tray': current_tray,
                'name': clean_name,
                'quantity': qty,
                'notes': f'{sheet_label}: {note}' if note else sheet_label,
            })
    return records


def load_rows(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def main():
    parser = argparse.ArgumentParser(description='Regenerate seed_search_products.dart from the two inventory registers.')
    parser.add_argument('--admin', default='cda_admin.xlsx', help='Path to the Branch 1 workbook (default: cda_admin.xlsx)')
    parser.add_argument('--ops', default='cda_ops.xlsx', help='Path to the Branch 2 workbook (default: cda_ops.xlsx)')
    parser.add_argument('--out', default='seed_search_products.dart', help='Output .dart file (default: seed_search_products.dart)')
    parser.add_argument('--json-out', default=None, help='Optional: also dump the parsed records as JSON for inspection')
    args = parser.parse_args()

    all_records = []

    # ── Branch 1 — cda_admin.xlsx, Sheet1 ──────────────────────────────
    wb1 = openpyxl.load_workbook(args.admin, data_only=True)
    ws1 = wb1['Sheet1']
    rows1 = load_rows(ws1)
    header1 = rows1[0]
    branch_col = header1.index('BRANCH')
    recs = parse_paired_or_unpaired_sheet(
        rows1, 0, branch='Branch 1', branch_col=branch_col, sheet_label='CDA Admin'
    )
    all_records.extend(recs)
    print(f'Branch 1 ({args.admin}): {len(recs)} records')

    # ── Branch 2 — cda_ops.xlsx, multiple sheets ───────────────────────
    wb2 = openpyxl.load_workbook(args.ops, data_only=True)

    # Sheet1: general rooms, one column per room (some with 3-col merge
    # spacer, values are DATE objects for the 'SERVICE & DELIVERY IN'
    # section which we skip since they aren't products).
    ws = wb2['Sheet1']
    rows = load_rows(ws)
    header = rows[0]
    # Drop columns whose header is a datetime (merged date columns) so
    # classify_columns doesn't choke on them, and 'GOJAN IN PRODUCTS'
    # date-pair columns (kept as room, its DATE IN/DATE OUT neighbors are
    # skipped automatically since they're literal 'DATE IN'/'DATE OUT').
    clean_header = []
    for h in header:
        if h is not None and not isinstance(h, str):
            clean_header.append(None)
        else:
            clean_header.append(h)
    rows[0] = clean_header
    recs = parse_paired_or_unpaired_sheet(rows, 0, branch='Branch 2', sheet_label='CDA Ops · Sheet1')
    all_records.extend(recs)
    print(f'Branch 2 Sheet1: {len(recs)} records')

    # STORAGE FACILITY
    ws = wb2['STORAGE FACILITY']
    rows = load_rows(ws)
    recs = parse_paired_or_unpaired_sheet(rows, 0, branch='Branch 2',
                                           extra_room_prefix='Storage Facility',
                                           sheet_label='CDA Ops · Storage Facility')
    all_records.extend(recs)
    print(f'Branch 2 Storage Facility: {len(recs)} records')

    # STORE ROOM
    ws = wb2['STORE ROOM']
    rows = load_rows(ws)
    recs = parse_paired_or_unpaired_sheet(rows, 0, branch='Branch 2', sheet_label='CDA Ops · Store Room')
    all_records.extend(recs)
    print(f'Branch 2 Store Room: {len(recs)} records')

    # R&D AND NEW PRODUCTS BERO
    ws = wb2['R&D AND NEW PRODUCTS BERO']
    rows = load_rows(ws)
    recs = parse_paired_or_unpaired_sheet(rows, 0, branch='Branch 2',
                                           extra_room_prefix='R&D Bero',
                                           sheet_label='CDA Ops · R&D Bero')
    all_records.extend(recs)
    print(f'Branch 2 R&D Bero: {len(recs)} records')

    # CHARGING STATION  — real header is row index 1 ('ITEMS'/'QUANTITES');
    # row index 0 ('UPPER ROW', None) is itself the first row-marker, so
    # splice it back in as the first data row under that header.
    ws = wb2['CHARGING STATION ']
    rows = load_rows(ws)
    spliced = [rows[1]] + [rows[0]] + rows[2:]
    recs = parse_paired_or_unpaired_sheet(spliced, 0, branch='Branch 2',
                                           extra_room_prefix='Charging Station',
                                           sheet_label='CDA Ops · Charging Station')
    # 'ITEMS' header would name the room "Items" — force the intended name.
    for r in recs:
        r['room'] = 'Charging Station'
    all_records.extend(recs)
    print(f'Branch 2 Charging Station: {len(recs)} records')

    # RPTO BERO
    ws = wb2['RPTO BERO']
    rows = load_rows(ws)
    recs = parse_paired_or_unpaired_sheet(rows, 0, branch='Branch 2',
                                           extra_room_prefix='RPTO Bero',
                                           sheet_label='CDA Ops · RPTO Bero')
    all_records.extend(recs)
    print(f'Branch 2 RPTO Bero: {len(recs)} records')

    print(f'\nTOTAL: {len(all_records)} records')

    if args.json_out:
        with open(args.json_out, 'w', encoding='utf-8') as f:
            json.dump(all_records, f, indent=1)
        print(f'Wrote {args.json_out}')

    write_dart_seed_file(all_records, args.out)
    print(f'Wrote {args.out}')
    print(f'Copy that file over lib/data/seed_search_products.dart in your Flutter project.')


# ═══════════════════════════════════════════════════════════════════════
#  DART FILE GENERATION
# ═══════════════════════════════════════════════════════════════════════
def _dart_escape(s: str) -> str:
    return s.replace('\\', '\\\\').replace("'", "\\'")


def _rec_line(r: dict) -> str:
    parts = [f"'name': '{_dart_escape(r['name'].strip())}'",
             f"'branch': '{r['branch']}'",
             f"'room': '{_dart_escape(r['room'])}'"]
    if r['row']:
        parts.append(f"'row': '{_dart_escape(r['row'])}'")
    if r['tray']:
        parts.append(f"'tray': '{_dart_escape(r['tray'])}'")
    parts.append(f"'quantity': {r['quantity']}")
    parts.append("'price': 0.0")
    if r.get('notes'):
        parts.append(f"'notes': '{_dart_escape(r['notes'])}'")
    return '    {' + ', '.join(parts) + '},'


def write_dart_seed_file(records: list, out_path: str) -> None:
    from collections import OrderedDict
    groups = OrderedDict()
    for r in records:
        groups.setdefault((r['branch'], r['room']), []).append(r)

    out = []
    out.append('// lib/data/seed_search_products.dart')
    out.append('//')
    out.append('// Full physical inventory seed for the Search Products screen, parsed')
    out.append('// programmatically from the two master registers (cda_admin.xlsx ->')
    out.append('// Branch 1, cda_ops.xlsx -> Branch 2). Every item carries its full')
    out.append('// physical location: branch / room / row / tray.')
    out.append('//')
    out.append(f'// Total items: {len(records)}')
    out.append('//')
    out.append('// Regenerate: python3 parse_inventory.py --admin cda_admin.xlsx --ops cda_ops.xlsx --out seed_search_products.dart')
    out.append('')
    out.append('class SeedSearchProducts {')
    out.append('  SeedSearchProducts._();')
    out.append('')
    out.append('  static List<Map<String, dynamic>> get allItems => [')

    cur_branch = None
    for (branch, room), items in groups.items():
        if branch != cur_branch:
            cur_branch = branch
            label = 'CDA Admin' if branch == 'Branch 1' else 'CDA Ops'
            out.append('')
            out.append('    // ' + '=' * 70)
            out.append(f'    // {branch.upper()} -- {label}')
            out.append('    // ' + '=' * 70)
        out.append(f'    // -- {room} ({len(items)} items) --')
        for r in items:
            out.append(_rec_line(r))

    out.append('  ];')
    out.append('}')
    out.append('')

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(out))


if __name__ == '__main__':
    main()